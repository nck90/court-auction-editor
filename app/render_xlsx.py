#!/usr/bin/env python3
"""정규화 JSON → 최종 표 Excel(.xlsx) 렌더러.

PDF/HTML 렌더러와 같은 구조를 재현한다:
  - 섹션: [아파트], [연립주택/다세대/빌라], [단독주택,다가구주택],
          [상가/오피스텔,근린시설], [대지/임야/전답], [기타]
  - 각 섹션에 사건번호/물건번호/소재지/용도/감정가/비고 테이블
  - 소재지·용도는 entry 의 properties 수만큼 행 생성, 나머지 컬럼은 병합(rowspan)
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from render_final_notice import GROUP_ORDER, compact_common, format_entry, load_entries, parse_case_sort_key


HEADER = ["사건번호", "물건번호", "소재지 및 면적[㎡]", "용도", "감정평가액\n최저매각가격\n[단위 : 원]", "비고"]
COL_WIDTHS = [14, 7, 60, 16, 18, 24]

THIN = Side(border_style="thin", color="4E4E4E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="F3F3F3")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
TITLE_FONT = Font(name="Malgun Gothic", size=16, bold=True)
GROUP_FONT = Font(name="Malgun Gothic", size=12, bold=True)
BASE_FONT = Font(name="Malgun Gothic", size=10)
HEADER_FONT = Font(name="Malgun Gothic", size=10, bold=True)


def _apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER


def render_xlsx(doc: dict, out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "매각공고"

    # 열 너비
    for idx, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    row = 1
    # 타이틀
    ws.cell(row=row, column=1, value="법원 경매부동산의 매각 공고").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADER))
    ws.row_dimensions[row].height = 26
    row += 1

    # 메타
    meta_lines = [
        "법원 경매부동산의 매각 공고",
        f"1.매각물건의 표시 및 매각조건 <{doc.get('court_line') or ''}>",
        doc.get("auction_datetime") or "",
        doc.get("decision_datetime") or "",
        doc.get("officer_line") or "",
    ]
    for line in meta_lines:
        if not line:
            continue
        c = ws.cell(row=row, column=1, value=line)
        c.font = BASE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADER))
        row += 1
    row += 1  # blank spacer

    # 엔트리 분류
    rendered = [
        format_entry(e)
        for e in doc.get("entries", [])
        if compact_common(e.get("usage", "")) not in {"자동차", "선박", "건설기계", "항공기"}
    ]
    grouped: dict[str, list[dict]] = {k: [] for k in GROUP_ORDER}
    for entry in rendered:
        grouped[entry["group"]].append(entry)

    for group in GROUP_ORDER:
        rows = grouped.get(group) or []
        if not rows:
            continue
        rows.sort(key=lambda r: parse_case_sort_key(r["case"], r["item"]))

        # 그룹 헤더
        gc = ws.cell(row=row, column=1, value=f"[{group}]")
        gc.font = GROUP_FONT
        gc.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADER))
        row += 1

        # 테이블 헤더
        for col_idx, label in enumerate(HEADER, 1):
            c = ws.cell(row=row, column=col_idx, value=label)
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.fill = HEADER_FILL
        _apply_border_range(ws, row, row, 1, len(HEADER))
        ws.row_dimensions[row].height = 36
        row += 1

        # 데이터 행
        for entry in rows:
            locations = list(entry.get("locations") or []) or [""]
            usages = list(entry.get("usages") or [])
            if len(usages) < len(locations):
                usages += [""] * (len(locations) - len(usages))
            span = len(locations)
            start_row = row
            end_row = row + span - 1

            # 사건번호/물건번호/감정가/비고 (span merge)
            ws.cell(row=start_row, column=1, value=entry.get("case", "")).alignment = LEFT
            ws.cell(row=start_row, column=2, value=entry.get("item", "")).alignment = CENTER
            ws.cell(row=start_row, column=5, value=entry.get("price", "")).alignment = LEFT
            ws.cell(row=start_row, column=6, value=entry.get("note", "")).alignment = LEFT
            for col in (1, 2, 5, 6):
                ws.cell(row=start_row, column=col).font = BASE_FONT
            if span > 1:
                for col in (1, 2, 5, 6):
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

            # 소재지/용도
            for offset, (loc, usage) in enumerate(zip(locations, usages)):
                r = start_row + offset
                c_loc = ws.cell(row=r, column=3, value=loc)
                c_loc.alignment = LEFT
                c_loc.font = BASE_FONT
                c_usage = ws.cell(row=r, column=4, value=usage)
                c_usage.alignment = LEFT
                c_usage.font = BASE_FONT

            _apply_border_range(ws, start_row, end_row, 1, len(HEADER))
            for r in range(start_row, end_row + 1):
                ws.row_dimensions[r].height = 24
            row = end_row + 1

        row += 1  # 섹션 간 공백

    # Freeze top (title + meta)
    ws.freeze_panes = "A7"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser(description="정규화 JSON을 Excel 최종본으로 렌더링")
    parser.add_argument("json_path", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    args = parser.parse_args()

    doc = load_entries(args.json_path)
    out = args.output or args.json_path.with_suffix(".xlsx")
    render_xlsx(doc, out)
    print(f"XLSX: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
