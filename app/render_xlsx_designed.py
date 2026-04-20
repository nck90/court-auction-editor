#!/usr/bin/env python3
"""정규화 JSON → A3 가로 2-컬럼 레이아웃 Excel(.xlsx) 렌더러.

InDesign 6단/5단 landscape 최종 공고지의 배치 개념을 보존한다:
  - 제목 / 메타 / 좌·우 2-컬럼 테이블 스택 / 하단 서명부
  - 좌측 컬럼 그룹: 단독주택,다가구주택 · 상가/오피스텔,근린시설 · 연립주택/다세대/빌라
  - 우측 컬럼 그룹: 아파트 · 기타 · 대지/임야/전답

각 그룹 테이블:
  사건번호 | 물건번호 | 소재지 및 면적[㎡] | 용도 | 감정평가액/최저매각가격 | 비고

엔트리당 properties(소재지/용도 쌍) 수만큼 행을 확장하고 사건번호·물건번호·감정가·비고는
해당 range 만큼 세로 병합한다.

Usage::

    from render_xlsx_designed import render_xlsx_designed
    render_xlsx_designed(doc_dict, Path("out.xlsx"))
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from render_final_notice import (
    GROUP_ORDER,
    compact_common,
    format_entry,
    load_entries,
    parse_case_sort_key,
)


# ---------------------------------------------------------------------------
# 컬럼 설계
# ---------------------------------------------------------------------------
# 좌측 컬럼(A~F): 사건번호/물건번호/소재지/용도/감정가/비고
# 간격 G
# 우측 컬럼(H~M): 사건번호/물건번호/소재지/용도/감정가/비고
LEFT_COLS = list(range(1, 7))       # A..F
GAP_COL = 7                          # G
RIGHT_COLS = list(range(8, 14))     # H..M
TOTAL_COLS = 13

COL_WIDTHS = {
    1: 11,   # 사건번호
    2: 5,    # 물건번호
    3: 44,   # 소재지
    4: 11,   # 용도
    5: 15,   # 감정가
    6: 16,   # 비고
    7: 2,    # 간격
    8: 11,
    9: 5,
    10: 44,
    11: 11,
    12: 15,
    13: 16,
}

HEADER_LABELS = [
    "사건번호",
    "물건번호",
    "소재지 및 면적[㎡]",
    "용도",
    "감정평가액\n최저매각가격\n[단위 : 원]",
    "비고",
]

LEFT_GROUPS = [
    "단독주택,다가구주택",
    "상가/오피스텔,근린시설",
    "연립주택/다세대/빌라",
]
RIGHT_GROUPS = [
    "아파트",
    "기타",
    "대지/임야/전답",
]


# ---------------------------------------------------------------------------
# 스타일 토큰
# ---------------------------------------------------------------------------
THIN = Side(border_style="thin", color="4E4E4E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="F3F3F3")
GROUP_FILL = PatternFill("solid", fgColor="FAF3E8")

FONT_TITLE = Font(name="Malgun Gothic", size=20, bold=True)
FONT_SUBTITLE = Font(name="Malgun Gothic", size=11, bold=True)
FONT_META = Font(name="Malgun Gothic", size=10)
FONT_GROUP = Font(name="Malgun Gothic", size=11, bold=True)
FONT_HEADER = Font(name="Malgun Gothic", size=10, bold=True)
FONT_BASE = Font(name="Malgun Gothic", size=9)
FONT_SIGN_BIG = Font(name="Malgun Gothic", size=14, bold=True)
FONT_SIGN = Font(name="Malgun Gothic", size=11)
FONT_NOTICE = Font(name="Malgun Gothic", size=8)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_CENTER = Alignment(horizontal="right", vertical="center", wrap_text=True)


# 하단에 붙는 법적 안내 템플릿(원 PDF 의 우측/하단 문구 요약).
LEGAL_NOTICE_LINES = [
    "2. 매각조건 및 유의사항",
    " 가. 매각 및 매각결정은 위에 기재된 일시·장소에서 행한다.",
    " 나. 매각방법 : 기일입찰",
    " 다. 매각장소 : 해당 법원 경매법정",
    " 라. 최저매각가격의 10% 해당액을 매수신청보증금으로 제공하여야 한다.",
    " 마. 매각물건명세서·현황조사보고서·감정평가서는 매각기일 1주일 전부터 법원에 비치한다.",
    " 바. 매각물건에 대한 권리관계, 임차인 현황 및 각종 특별매각조건은 매각물건명세서를 통해 확인하여야 한다.",
    " 사. 대금지급기한은 매각허가결정이 확정된 날부터 통상 1개월 이내이며, 기한 내 미납 시 차순위매수신고인 또는 재매각 절차가 진행된다.",
    " 아. 공유자 우선매수권 등 특별매각조건이 있는 경우 해당 물건 비고란에 표시되며, 관련 법령 및 절차를 따른다.",
]


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _apply_border_range(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER


def _write_cell(ws: Worksheet, row: int, col: int, value: str, *, font=FONT_BASE, alignment=LEFT_TOP, fill=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = alignment
    if fill is not None:
        c.fill = fill


def _estimate_row_height(text: str, col_width: float, base: float = 16.0, per_line: float = 14.0) -> float:
    """셀 텍스트 높이를 대략 추정한다(wrap 고려)."""
    if not text:
        return base
    max_lines = 0
    for line in str(text).splitlines() or [""]:
        # 한글은 대충 컬럼 너비의 1.6배 정도 글자를 수용한다고 가정.
        capacity = max(int(col_width * 1.6), 1)
        n = max(1, (len(line) + capacity - 1) // capacity)
        max_lines = max(max_lines, n)
    max_lines = max(max_lines, len(str(text).splitlines()) or 1)
    return max(base, per_line * max_lines + 4)


def _classify_entries(doc: dict) -> dict[str, list[dict]]:
    rendered = [
        format_entry(e)
        for e in doc.get("entries", [])
        if compact_common(e.get("usage", "")) not in {"자동차", "선박", "건설기계", "항공기"}
    ]
    buckets: dict[str, list[dict]] = {k: [] for k in GROUP_ORDER}
    for entry in rendered:
        buckets.setdefault(entry["group"], []).append(entry)
    for rows in buckets.values():
        rows.sort(key=lambda r: parse_case_sort_key(r["case"], r["item"]))
    return buckets


def _render_header_row(ws: Worksheet, row: int, cols: list[int]) -> None:
    for offset, col in enumerate(cols):
        _write_cell(
            ws,
            row,
            col,
            HEADER_LABELS[offset],
            font=FONT_HEADER,
            alignment=CENTER,
            fill=HEADER_FILL,
        )
    _apply_border_range(ws, row, row, cols[0], cols[-1])
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 42)


def _render_group_header(ws: Worksheet, row: int, cols: list[int], group: str) -> None:
    c = ws.cell(row=row, column=cols[0], value=f"[{group}]")
    c.font = FONT_GROUP
    c.alignment = LEFT_CENTER
    c.fill = GROUP_FILL
    ws.merge_cells(start_row=row, start_column=cols[0], end_row=row, end_column=cols[-1])
    # border on the whole merged range
    _apply_border_range(ws, row, row, cols[0], cols[-1])
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 22)


def _render_entry(ws: Worksheet, row: int, cols: list[int], entry: dict) -> int:
    """하나의 엔트리를 row 부터 써넣고 다음 free row 번호를 돌려준다."""
    locations = list(entry.get("locations") or []) or [""]
    usages = list(entry.get("usages") or [])
    if len(usages) < len(locations):
        usages += [""] * (len(locations) - len(usages))
    span = len(locations)
    start_row = row
    end_row = row + span - 1

    col_case, col_item, col_loc, col_usage, col_price, col_note = cols

    _write_cell(ws, start_row, col_case, entry.get("case", ""), alignment=CENTER)
    _write_cell(ws, start_row, col_item, entry.get("item", ""), alignment=CENTER)
    _write_cell(ws, start_row, col_price, entry.get("price", ""), alignment=CENTER)
    _write_cell(ws, start_row, col_note, entry.get("note", ""), alignment=LEFT_TOP)

    if span > 1:
        for col in (col_case, col_item, col_price, col_note):
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=end_row,
                end_column=col,
            )

    loc_width = COL_WIDTHS.get(col_loc, 44)
    for offset, (loc, usage) in enumerate(zip(locations, usages)):
        r = start_row + offset
        _write_cell(ws, r, col_loc, loc, alignment=LEFT_TOP)
        _write_cell(ws, r, col_usage, usage, alignment=CENTER)
        # 행 높이: 소재지와 비고 중 더 긴 쪽 기준.
        height = _estimate_row_height(loc, loc_width, base=18.0, per_line=14.0)
        if offset == 0:
            # 비고/가격도 포함해 첫 행 높이 잡기(병합이라 첫 행에 모인다).
            note_h = _estimate_row_height(entry.get("note", ""), COL_WIDTHS.get(col_note, 16), base=18.0, per_line=14.0)
            price_h = _estimate_row_height(entry.get("price", ""), COL_WIDTHS.get(col_price, 15), base=18.0, per_line=14.0)
            height = max(height, note_h / max(span, 1), price_h / max(span, 1))
        current = ws.row_dimensions[r].height or 0
        ws.row_dimensions[r].height = max(current, height)

    _apply_border_range(ws, start_row, end_row, cols[0], cols[-1])
    return end_row + 1


def _render_column(ws: Worksheet, start_row: int, cols: list[int], groups: list[str], buckets: dict[str, list[dict]]) -> int:
    row = start_row
    first = True
    for group in groups:
        rows = buckets.get(group) or []
        if not rows:
            continue
        if not first:
            row += 1  # 섹션 사이 공백
        first = False
        _render_group_header(ws, row, cols, group)
        row += 1
        _render_header_row(ws, row, cols)
        row += 1
        for entry in rows:
            row = _render_entry(ws, row, cols, entry)
    return row


# ---------------------------------------------------------------------------
# 퍼블릭 엔트리 포인트
# ---------------------------------------------------------------------------

def render_xlsx_designed(doc: dict, out_path: Path) -> Path:
    """A3 가로 2-컬럼 레이아웃 xlsx 를 생성해 out_path 에 저장한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "매각공고"

    # 페이지 설정: A3 landscape, fit to width
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # 열 너비
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 1행: 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title = ws.cell(row=1, column=1, value="법원 경매부동산의 매각 공고")
    title.font = FONT_TITLE
    title.alignment = CENTER
    ws.row_dimensions[1].height = 34

    # 2행: 소제목
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(row=2, column=1, value="1. 매각물건의 표시 및 매각조건")
    sub.font = FONT_SUBTITLE
    sub.alignment = LEFT_CENTER
    ws.row_dimensions[2].height = 22

    # 3~6행: 메타 정보 (좌/우 분리 가능하나 우선 좌측 전폭)
    court_line = doc.get("court_line") or ""
    auction_dt = doc.get("auction_datetime") or ""
    decision_dt = doc.get("decision_datetime") or ""
    officer = doc.get("officer_line") or ""

    meta_row = 3
    # court_line: 예) "담 당 계 : 경매2계 김학상"
    ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=TOTAL_COLS)
    c = ws.cell(row=meta_row, column=1, value=f"<{court_line}>" if court_line else "")
    c.font = FONT_SUBTITLE
    c.alignment = LEFT_CENTER
    ws.row_dimensions[meta_row].height = 20
    meta_row += 1

    for line in (auction_dt, decision_dt, officer):
        if not line:
            continue
        ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=TOTAL_COLS)
        c = ws.cell(row=meta_row, column=1, value=line)
        c.font = FONT_META
        c.alignment = LEFT_CENTER
        ws.row_dimensions[meta_row].height = 18
        meta_row += 1

    # 공백 행
    meta_row += 1
    table_start = meta_row

    # 엔트리 분류
    buckets = _classify_entries(doc)

    # 좌/우 컬럼 테이블 스택을 같은 row offset 부터 시작.
    left_end = _render_column(ws, table_start, LEFT_COLS, LEFT_GROUPS, buckets)
    right_end = _render_column(ws, table_start, RIGHT_COLS, RIGHT_GROUPS, buckets)

    body_end = max(left_end, right_end)

    # 하단 안내
    footer_row = body_end + 1
    for line in LEGAL_NOTICE_LINES:
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=TOTAL_COLS)
        c = ws.cell(row=footer_row, column=1, value=line)
        c.font = FONT_NOTICE
        c.alignment = LEFT_CENTER
        ws.row_dimensions[footer_row].height = 16
        footer_row += 1

    footer_row += 1
    # 날짜·법원·서명
    date_line = _extract_posting_date(doc)
    court_name, judge_name = _extract_court_and_judge(doc)

    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=TOTAL_COLS)
    c = ws.cell(row=footer_row, column=1, value=date_line)
    c.font = FONT_SIGN
    c.alignment = RIGHT_CENTER
    ws.row_dimensions[footer_row].height = 20
    footer_row += 1

    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=TOTAL_COLS)
    c = ws.cell(row=footer_row, column=1, value=court_name)
    c.font = FONT_SIGN_BIG
    c.alignment = CENTER
    ws.row_dimensions[footer_row].height = 28
    footer_row += 1

    if judge_name:
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=TOTAL_COLS)
        c = ws.cell(row=footer_row, column=1, value=judge_name)
        c.font = FONT_SIGN
        c.alignment = RIGHT_CENTER
        ws.row_dimensions[footer_row].height = 22
        footer_row += 1

    # 인쇄 영역
    end_col = get_column_letter(TOTAL_COLS)
    ws.print_area = f"A1:{end_col}{footer_row - 1}"
    ws.freeze_panes = f"A{table_start}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _extract_posting_date(doc: dict) -> str:
    """auction_datetime 에서 날짜만 추출해 '2026. 2. 6.' 같은 형태로."""
    import re

    source = doc.get("auction_datetime") or doc.get("decision_datetime") or ""
    m = re.search(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\.", source)
    if m:
        return f"{m.group(1)}.  {int(m.group(2))}.  {int(m.group(3))}."
    return ""


def _extract_court_and_judge(doc: dict) -> tuple[str, str]:
    """court_line("담 당 계 : 경매2계 김학상") 에서 담당자 성명을 찾아 반환."""
    import re

    court_line = doc.get("court_line") or ""
    judge = ""
    officer = doc.get("officer_line") or ""

    # 공고지 하단: "법원명"이 필요하지만 정규화 JSON 에 별도 필드가 없다.
    # 대신 경매 계 + 담당자 이름을 활용해 표시.
    court_name = ""
    court_match = re.search(r"([가-힣]+(?:지방)?법원(?:\s+[가-힣]+지원)?)", officer) or re.search(r"([가-힣]+(?:지방)?법원(?:\s+[가-힣]+지원)?)", court_line)
    if court_match:
        court_name = court_match.group(1)

    judge_match = re.search(r"경매\d*계\s+([가-힣]{2,4})", court_line)
    if judge_match:
        judge = f"사법보좌관  {judge_match.group(1)}"
    elif officer:
        judge = officer

    if not court_name:
        court_name = "법원"
    return court_name, judge


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="정규화 JSON을 A3 가로 2-컬럼 디자인 Excel로 렌더링")
    parser.add_argument("json_path", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=None)
    args = parser.parse_args()

    doc = load_entries(args.json_path)
    out = args.output or args.json_path.with_name(f"{args.json_path.stem}.디자인.xlsx")
    render_xlsx_designed(doc, out)
    print(f"XLSX: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
