"""Render the final grouped notice as a styled XLSX workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
GROUP_ORDER = [
    "아파트",
    "연립주택/다세대/빌라",
    "단독주택,다가구주택",
    "상가/오피스텔,근린시설",
    "대지/임야/전답",
    "기타",
]


def _set_border(ws, row: int, col_start: int, col_end: int, border: Border) -> None:
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = border


def render_xlsx(final_json: dict, xlsx_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "매각공고"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 18,
        "B": 9,
        "C": 50,
        "D": 14,
        "E": 24,
        "F": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    title_font = Font(name="Malgun Gothic", size=14, bold=True)
    head_font = Font(name="Malgun Gothic", size=10, bold=True)
    body_font = Font(name="Malgun Gothic", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="법원 경매부동산의 매각 공고")
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[row].height = 26
    row += 2

    header = final_json.get("header", {}) or {}
    info_rows = [
        ("담 당 계", header.get("damdang", "")),
        ("매각일시", header.get("sale_date", "")),
        ("매각결정일시", header.get("decision_date", "")),
        ("장소", header.get("location", "")),
    ]
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = head_font
        ws.cell(row=row, column=2, value=value).font = body_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2).alignment = left
        row += 1
    row += 1

    records = final_json.get("records", []) or []
    grouped = {group: [] for group in GROUP_ORDER}
    for rec in records:
        group = (rec.get("group") or "기타").strip()
        if group in grouped:
            grouped[group].append(rec)

    headers = [
        "사건번호",
        "물건번호",
        "소재지 및 면적[㎡]",
        "용도",
        "감정평가액\n최저매각가격\n[단위:원]",
        "비고",
    ]

    for group in GROUP_ORDER:
        items = grouped.get(group) or []
        if not items:
            continue

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=f"[{group}]")
        c.font = Font(name="Malgun Gothic", size=11, bold=True)
        c.alignment = left
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value="(단위 : 원)")
        c.font = body_font
        c.alignment = right
        row += 1

        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = head_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[row].height = 32
        row += 1

        for rec in items:
            locs = rec.get("locations") or [{"address": "", "use": ""}]
            start_row = row
            for loc in locs:
                ws.cell(row=row, column=3, value=loc.get("address", "")).font = body_font
                ws.cell(row=row, column=3).alignment = left
                ws.cell(row=row, column=4, value=loc.get("use", "")).font = body_font
                ws.cell(row=row, column=4).alignment = center
                _set_border(ws, row, 1, 6, border)
                row += 1

            end_row = row - 1
            if end_row > start_row:
                for col in (1, 2, 5, 6):
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col,
                        end_row=end_row,
                        end_column=col,
                    )

            ws.cell(start_row, column=1, value=rec.get("case_no", "")).font = body_font
            if rec.get("dup_tag"):
                ws.cell(start_row, column=1).value = f"{rec.get('case_no', '')}\n{rec.get('dup_tag', '')}"
            ws.cell(start_row, column=1).alignment = center

            ws.cell(start_row, column=2, value=rec.get("item_no", "")).font = body_font
            ws.cell(start_row, column=2).alignment = center

            ws.cell(
                start_row,
                column=5,
                value=f"{rec.get('price', '')}\n{rec.get('min_price', '')}",
            ).font = body_font
            ws.cell(start_row, column=5).alignment = right

            ws.cell(start_row, column=6, value=rec.get("note", "")).font = body_font
            ws.cell(start_row, column=6).alignment = left

        row += 1

    out = Path(xlsx_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
